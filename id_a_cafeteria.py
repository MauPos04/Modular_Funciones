import pandas as pd
from openpyxl import load_workbook
import os
from Eiminar_columnas_yAmarillo import eliminar_columnas_y_colorear   

def renombrar_hojas():
    # Diccionario de correspondencia entre ID y nombre
    id_a_nombre = {
        7: 'Zamari Gourmet - Sede Zúñiga',
        8: 'Nativos - ECI ',
        10: 'Dogger - UPB',
        3: 'Coco Melón - CES Poblado',
        2: 'Dulcinea - CES Poblado',
        21: 'BOTY',
        14: 'Pimientoz UPB',
        18: 'El Buñuelo - CES',
        25: 'FUNDACIÓN PUTCHI',
        12: 'Taco Factory - UPB',
        9: 'Aldea Nikkei - EAFIT',
        27: 'Tradiciones Madecentro',
        4: 'Zamari Gourmet - Sede Palmas',
        26: 'Chip Station',
        13: 'ServiExpress',
        16: 'Pizzotas',
        17: 'El Buñuelo - EIA',
        6: 'Dogger - EIA',
        20: 'Montana',
        19: 'COLOSAL PETS',
        1: 'Feroz Helados de Yogurt - UPB',
        0: 'De Lolita - EIA',
        5: 'La Cafetería - Idiomas',
        22: 'Nativos - EIA',
        24: 'EINSTEIN',
        23: 'Tradiciones Culinarias',
        15: 'Pimientoz',
        11: 'Taco Factory - EAFIT'
    }

    # Ruta del archivo Excel
    ruta_archivo = r'excel_exports\Ordenes_Completadas_Por_Cafeteria.xlsx'  # Aquí pon la ruta correcta a tu archivo

    # Cargar el archivo Excel con pandas para manipular las fechas
    df = pd.read_excel(ruta_archivo, sheet_name=None)  # Lee todas las hojas del archivo

    # Recorremos cada hoja para filtrar las fechas
    for sheet_name, data in df.items():
        # Verificar si la columna 'fecha_creacion_str' existe en la hoja
        if 'fecha_creacion_str' in data.columns:
            # Convertir la columna 'fecha_creacion_str' a tipo datetime
            data['fecha_creacion'] = pd.to_datetime(data['fecha_creacion_str'], format='%d/%m/%Y')
            
            # Definir la fecha de corte: 18-10-2024
            fecha_corte = pd.to_datetime('18/10/2024', format='%d/%m/%Y')
            
            # Filtrar las filas donde la fecha es mayor o igual a la fecha de corte
            data = data[data['fecha_creacion'] >= fecha_corte]
            
            # Eliminar la columna temporal 'fecha_creacion' utilizada para el filtrado
            data = data.drop(columns=['fecha_creacion'])
            
            # Guardar los datos filtrados de nuevo en el diccionario
            df[sheet_name] = data
    
    # Guardar el archivo Excel con los cambios de fecha y formato de tabla
    with pd.ExcelWriter(ruta_archivo, engine='xlsxwriter') as writer:
        for sheet_name, data in df.items():
            # Escribir cada hoja con los datos filtrados
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            num_rows, num_cols = data.shape
            
            # Aplicar formato de tabla
            worksheet.add_table(0, 0, num_rows, num_cols - 1, {'columns': [{'header': col} for col in data.columns], 'name': 'Tabla' + sheet_name, 'style': 'Table Style Medium 9'})

    # Cargar el archivo Excel con openpyxl para renombrar las hojas
    wb = load_workbook(ruta_archivo)

    # Obtener los nombres actuales de las hojas
    hojas_actuales = wb.sheetnames
    print("Hojas actuales:", hojas_actuales)

    # Renombrar las hojas en el libro de trabajo
    for hoja in hojas_actuales:
        try:
            # Extraemos el número de la hoja (que debería ser su identificador)
            sheet_id = int(hoja)
            
            # Verificamos si ese ID existe en el diccionario
            if sheet_id in id_a_nombre:
                # Cambiar el nombre de la hoja
                sheet = wb[hoja]
                sheet.title = id_a_nombre[sheet_id]
        except ValueError:
            # Si el nombre de la hoja no es un número (no esperado), ignoramos esta hoja
            continue

    # Guardar el archivo con los nuevos nombres de las hojas
    wb.save(ruta_archivo)  # Guardar sobre el archivo original, o puedes poner una ruta diferente

if __name__ == "__main__":
    renombrar_hojas()
    
